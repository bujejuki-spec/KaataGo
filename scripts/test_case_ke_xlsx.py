"""Mengubah TEST-CASE-KAATAGO.md jadi berkas Excel yang bisa dipakai menguji.

Test case bukan bahan bacaan seperti FSD dan TSD — ia daftar kerja.
Penguji perlu menyaring per prioritas, mengurutkan per modul, dan
menuliskan hasilnya di sebelah tiap baris. Dokumen Word tidak bisa
melakukan satu pun dari itu.

Karena itu keluarannya menambahkan tiga kolom yang tidak ada di
markdown-nya: Hasil, Catatan, dan Diuji Oleh. Kolom Hasil punya daftar
pilihan (Lulus / Gagal / Terblokir / Dilewati) supaya isinya seragam —
"OK", "ok", dan "Oke" di tiga baris berbeda membuat rekapnya harus
dibaca satu per satu.

Pakai:
    /usr/bin/python3 scripts/test_case_ke_xlsx.py
"""

import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SUMBER = os.path.join(REPO, "docs", "TEST-CASE-KAATAGO.md")
TUJUAN = os.path.join(REPO, "docs", "TEST-CASE-KAATAGO.xlsx")

UNGU = "4F46E5"
UNGU_MUDA = "EEEDFC"
ABU = "F1F5F9"
PUTIH = "FFFFFF"
MERAH = "B91C1C"
KUNING = "B45309"
HIJAU = "047857"

WARNA_P = {"P1": MERAH, "P2": KUNING, "P3": HIJAU}

tepi = Side(style="thin", color="D8DBE8")
KOTAK = Border(left=tepi, right=tepi, top=tepi, bottom=tepi)


def bersih(teks):
    """Membuang penanda markdown supaya selnya terbaca sebagai kalimat."""
    teks = re.sub(r"\*\*(.+?)\*\*", r"\1", teks)
    teks = re.sub(r"`(.+?)`", r"\1", teks)
    return teks.strip()


def baca(path):
    """Membaca tiap tabel kasus uji berikut bab tempatnya berada.

    Yang diambil hanya tabel berkolom lima yang barisnya dimulai TC-.
    Tabel lain di dokumen itu — prioritas, prasyarat, keterlacakan —
    memang bukan kasus uji, dan menyeretnya ikut masuk akan membuat
    penghitungan di kolom rekap salah tanpa terlihat salah.
    """
    bab = None
    hasil = []
    for baris in open(path).read().split("\n"):
        if baris.startswith("## "):
            bab = re.sub(r"^##\s+\d+\.\s*", "", baris).strip()
        if not baris.startswith("| TC-"):
            continue
        sel = [bersih(s) for s in baris.strip().strip("|").split(" | ")]
        if len(sel) != 5:
            continue
        tc, p, skenario, harapan, rujukan = sel
        hasil.append((bab, tc, p, skenario, harapan, rujukan))
    return hasil


def e2e(path):
    """Alur ujung-ke-ujung ditulis sebagai daftar bernomor, bukan tabel.

    Langkahnya digabung jadi satu sel bermultibaris — dipecah jadi satu
    baris per langkah akan memberi kesan tiap langkah punya hasilnya
    sendiri, padahal yang dinilai adalah alurnya sebagai satu kesatuan.
    """
    isi = open(path).read()
    blok = re.findall(r"### (E2E-\d+) — (.+?) \((P\d)\)\n(.*?)(?=\n### |\n---)",
                      isi, re.S)
    hasil = []
    for kode, judul, p, badan in blok:
        langkah = [bersih(b) for b in re.findall(r"^\d+\. (.+)$", badan, re.M)]
        if not langkah:
            langkah = [bersih(b) for b in badan.strip().split("\n") if b.strip()
                       and not b.startswith("**")]
        rujuk = re.search(r"\*\*Rujukan:\*\* (.+)", badan)
        hasil.append((
            "Uji Ujung-ke-Ujung", kode, p, bersih(judul),
            "\n".join(f"{i+1}. {l}" for i, l in enumerate(langkah)),
            bersih(rujuk.group(1)) if rujuk else "",
        ))
    return hasil


def judul_kolom(ws, kolom, lebar):
    for i, (nama, w) in enumerate(zip(kolom, lebar), start=1):
        sel = ws.cell(row=1, column=i, value=nama)
        sel.font = Font(bold=True, color=PUTIH, size=11)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.alignment = Alignment(vertical="center", horizontal="left",
                                  wrap_text=True)
        sel.border = KOTAK
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"


def lembar_kasus(wb, kasus):
    ws = wb.create_sheet("Kasus Uji")
    kolom = ["Modul", "ID", "P", "Skenario / Langkah",
             "Hasil yang Diharapkan", "Rujukan",
             "Hasil", "Catatan", "Diuji Oleh"]
    judul_kolom(ws, kolom, [22, 12, 6, 52, 52, 20, 13, 34, 16])

    for baris in kasus:
        ws.append(list(baris) + ["", "", ""])

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(kolom) + 1):
            sel = ws.cell(row=r, column=c)
            sel.border = KOTAK
            sel.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=3).font = Font(
            bold=True, color=WARNA_P.get(ws.cell(row=r, column=3).value, "000000"))
        ws.cell(row=r, column=3).alignment = Alignment(
            vertical="top", horizontal="center")
        ws.cell(row=r, column=2).font = Font(bold=True)
        if r % 2 == 0:
            for c in range(1, len(kolom) + 1):
                ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=ABU)

    # Daftar pilihan supaya isinya seragam. Tanpa ini, rekapnya harus
    # dibaca satu per satu untuk tahu "ok" dan "OK" itu hal yang sama.
    dv = DataValidation(
        type="list",
        formula1='"Lulus,Gagal,Terblokir,Dilewati"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"G2:G{ws.max_row}")

    ws.auto_filter.ref = f"A1:I{ws.max_row}"
    return ws


def lembar_rekap(wb, ws_kasus, kasus):
    ws = wb.create_sheet("Rekap", 0)
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    ws["A1"] = "KaataGo — Rekap Pengujian"
    ws["A1"].font = Font(bold=True, size=15, color=UNGU)
    ws["A2"] = "Angkanya menghitung sendiri dari lembar Kasus Uji."
    ws["A2"].font = Font(size=10, italic=True, color="6E728C")

    n = ws_kasus.max_row
    kepala = ["Kelompok", "Jumlah", "Lulus", "Gagal", "Terblokir",
              "Dilewati", "Belum diuji"]
    for i, teks in enumerate(kepala, start=1):
        sel = ws.cell(row=4, column=i, value=teks)
        sel.font = Font(bold=True, color=PUTIH)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.border = KOTAK
        sel.alignment = Alignment(horizontal="center")

    def rumus(baris, kriteria_kolom, kriteria, hasil=None):
        if hasil is None:
            return f'=COUNTIF(\'Kasus Uji\'!{kriteria_kolom}2:{kriteria_kolom}{n},"{kriteria}")'
        return (f'=COUNTIFS(\'Kasus Uji\'!{kriteria_kolom}2:{kriteria_kolom}{n},"{kriteria}",'
                f"'Kasus Uji'!G2:G{n},\"{hasil}\")")

    r = 5
    for p in ["P1", "P2", "P3"]:
        ws.cell(row=r, column=1, value=f"Prioritas {p}")
        ws.cell(row=r, column=2, value=rumus(r, "C", p))
        for i, h in enumerate(["Lulus", "Gagal", "Terblokir", "Dilewati"],
                              start=3):
            ws.cell(row=r, column=i, value=rumus(r, "C", p, h))
        ws.cell(row=r, column=7,
                value=f'=B{r}-SUM(C{r}:F{r})')
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, 8):
        L = get_column_letter(c)
        ws.cell(row=r, column=c, value=f"=SUM({L}5:{L}{r-1})").font = Font(bold=True)
    total_baris = r
    r += 2

    ws.cell(row=r, column=1, value="Per modul").font = Font(bold=True)
    r += 1
    for i, teks in enumerate(kepala, start=1):
        sel = ws.cell(row=r, column=i, value=teks)
        sel.font = Font(bold=True, color=PUTIH)
        sel.fill = PatternFill("solid", fgColor=UNGU)
        sel.border = KOTAK
        sel.alignment = Alignment(horizontal="center")
    r += 1

    for modul in dict.fromkeys(k[0] for k in kasus):
        ws.cell(row=r, column=1, value=modul)
        ws.cell(row=r, column=2, value=rumus(r, "A", modul))
        for i, h in enumerate(["Lulus", "Gagal", "Terblokir", "Dilewati"],
                              start=3):
            ws.cell(row=r, column=i, value=rumus(r, "A", modul, h))
        ws.cell(row=r, column=7, value=f"=B{r}-SUM(C{r}:F{r})")
        r += 1

    for baris in list(range(5, total_baris + 1)) + list(range(total_baris + 4, r)):
        for c in range(1, 8):
            sel = ws.cell(row=baris, column=c)
            sel.border = KOTAK
            if c > 1:
                sel.alignment = Alignment(horizontal="center")

    ws.cell(row=r + 1, column=1,
            value="Rilis ditahan selama masih ada P1 yang Gagal.").font = Font(
        italic=True, color=MERAH)
    return ws


def lembar_prasyarat(wb, path):
    ws = wb.create_sheet("Prasyarat")
    judul_kolom(ws, ["#", "Prasyarat", "Kenapa", "Siap?"], [8, 58, 62, 10])
    isi = open(path).read()
    for baris in isi.split("\n"):
        if not baris.startswith("| L-"):
            continue
        sel = [bersih(s) for s in baris.strip().strip("|").split(" | ")]
        if len(sel) == 3:
            ws.append(sel + [""])
    for r in range(2, ws.max_row + 1):
        for c in range(1, 5):
            k = ws.cell(row=r, column=c)
            k.border = KOTAK
            k.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=r, column=1).font = Font(bold=True)
    dv = DataValidation(type="list", formula1='"Ya,Belum"', allow_blank=True,
                        showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"D2:D{ws.max_row}")
    return ws


def main():
    kasus = baca(SUMBER) + e2e(SUMBER)

    wb = Workbook()
    wb.remove(wb.active)

    ws_kasus = lembar_kasus(wb, kasus)
    lembar_prasyarat(wb, SUMBER)
    lembar_rekap(wb, ws_kasus, kasus)

    wb.active = 0
    wb.save(TUJUAN)
    print(f"ditulis: {TUJUAN}")
    print(f"  {len(kasus)} kasus, {len(set(k[0] for k in kasus))} modul")


if __name__ == "__main__":
    main()
